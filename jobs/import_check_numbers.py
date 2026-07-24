"""
jobs/import_check_numbers.py — ΜΙΑΣ ΧΡΗΣΗΣ.

Μεταφέρει τους αριθμούς επιταγής από το Excel (ΜΗΝΑΣ.xlsx) στο φύλλο
«timologiseis» του Google Sheet, ταιριάζοντας με βάση την ΗΜΕΡΟΜΗΝΙΑ ΕΠΙΤΑΓΗΣ.

┌────────────────────────────────────────────────────────────────────────────┐
│ ΑΣΦΑΛΕΙΑ — ΤΙ ΚΑΝΕΙ ΚΑΙ ΤΙ ΔΕΝ ΚΑΝΕΙ                                       │
│                                                                            │
│ ✓ Γράφει ΜΟΝΟ σε γραμμές που έχουν ΑΔΕΙΟ αριθμό επιταγής.                 │
│ ✗ ΔΕΝ πειράζει ό,τι έχεις ήδη συμπληρώσει — η δουλειά σου είναι ιερή.      │
│ ✗ ΔΕΝ σβήνει, ΔΕΝ προσθέτει γραμμές. Μόνο συμπληρώνει κενά.                │
│                                                                            │
│ Τρέχει ΠΡΩΤΑ σε δοκιμαστική λειτουργία (--dry) και δείχνει τι θα κάνει.    │
└────────────────────────────────────────────────────────────────────────────┘

Χρήση:
    python jobs/import_check_numbers.py --file ΜΗΝΑΣ.xlsx --dry    # δοκιμή
    python jobs/import_check_numbers.py --file ΜΗΝΑΣ.xlsx          # εκτέλεση

Secrets: GOOGLE_KEY_JSON
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.sheets import load_timologiseis, update_timologiseis_field


# Πού κάθονται τα δεδομένα μέσα στο Excel.
# Τα φύλλα έχουν την ίδια δομή, αλλάζει μόνο η γραμμή της επικεφαλίδας.
SHEETS = {
    "2024 -2025": 3,
    "2026": 4,
    "2027": 3,
    "2028": 3,
}

COL_DATE = 2      # B — ΗΜΕΡ.
COL_AMOUNT = 3    # C — ΠΟΣΟ
COL_NUMBER = 4    # D — Αρ.Επιταγής


def read_excel(path: str) -> list[dict]:
    """Διαβάζει όλους τους αριθμούς επιταγής από όλα τα φύλλα."""
    try:
        import openpyxl
    except ImportError:
        print("✗ Λείπει το openpyxl. Τρέξε: pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    out = []

    for sheet, header_row in SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]

        for r in range(header_row + 1, ws.max_row + 1):
            d = ws.cell(r, COL_DATE).value
            num = ws.cell(r, COL_NUMBER).value

            if not d or not num:
                continue

            num = str(num).strip()
            if not num or num.lower() == "none":
                continue

            out.append({
                "date": d.date() if hasattr(d, "date") else d,
                "amount": float(ws.cell(r, COL_AMOUNT).value or 0),
                "number": num,
                "source": f"{sheet}!{r}",
            })

    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="Διαδρομή στο ΜΗΝΑΣ.xlsx")
    ap.add_argument("--dry", action="store_true", help="Δείξε τι θα γίνει, χωρίς εγγραφή")
    args = ap.parse_args()

    if not os.environ.get("GOOGLE_KEY_JSON"):
        print("✗ Λείπει το GOOGLE_KEY_JSON")
        return 1

    if not os.path.exists(args.file):
        print(f"✗ Δεν βρέθηκε το αρχείο: {args.file}")
        return 1

    # ── 1. ΔΙΑΒΑΣΕ ΤΟ EXCEL ──
    excel = read_excel(args.file)
    print(f"▶ Excel: {len(excel)} αριθμοί επιταγής")
    if not excel:
        return 1

    by_date = {r["date"]: r for r in excel}

    # ── 2. ΔΙΑΒΑΣΕ ΤΟ SHEET ──
    df = load_timologiseis()
    if df.empty:
        print("✗ Το φύλλο timologiseis είναι άδειο")
        return 1

    print(f"▶ Sheet: {len(df)} τιμολογήσεις")
    print()

    # ── 3. ΤΑΙΡΙΑΞΕ ──
    #
    # Πρώτα με ΑΚΡΙΒΗ ημερομηνία. Αν δεν βρεθεί, δοκιμάζουμε ±3 μέρες — γιατί
    # καμιά φορά η ίδια επιταγή είναι καταχωρημένη Δευτέρα στο ένα και Τρίτη
    # στο άλλο. Τα «κοντινά» ταιριάσματα επισημαίνονται ξεχωριστά, ώστε να τα
    # ελέγξεις με το μάτι πριν γραφτούν.
    to_write = []     # (row, number, date, amount)
    already = 0       # έχουν ήδη αριθμό — δεν τα πειράζουμε
    no_match = []     # δεν βρέθηκαν στο Excel
    used = set()      # αριθμοί που δόθηκαν ήδη — να μην μπει ο ίδιος 2 φορές

    from datetime import timedelta

    for _, r in df.iterrows():
        cd = r["check_date"]
        cd = cd.date() if hasattr(cd, "date") else cd
        existing = str(r.get("check_number", "") or "").strip()

        if existing:
            already += 1
            continue

        match = by_date.get(cd)
        near = False

        if not match:
            # Δοκίμασε ±3 μέρες, ξεκινώντας από την πιο κοντινή.
            for delta in (1, -1, 2, -2, 3, -3):
                cand = by_date.get(cd + timedelta(days=delta))
                if cand and cand["number"] not in used:
                    match, near = cand, True
                    break

        if not match or match["number"] in used:
            no_match.append(cd)
            continue

        used.add(match["number"])
        to_write.append({
            "row": int(r["_row"]),
            "number": match["number"],
            "date": cd,
            "excel_date": match["date"],
            "near": near,
            "sheet_amount": float(r["amount"]),
            "excel_amount": match["amount"],
        })

    # ── 4. ΑΝΑΦΟΡΑ ──
    exact = [w for w in to_write if not w["near"]]
    nears = [w for w in to_write if w["near"]]

    print(f"  Έχουν ήδη αριθμό (δεν πειράζονται): {already}")
    print(f"  Θα συμπληρωθούν:                    {len(to_write)}")
    print(f"      · ακριβής ημερομηνία:           {len(exact)}")
    print(f"      · κοντινή (±3 μέρες):           {len(nears)}")
    print(f"  Χωρίς αντιστοιχία στο Excel:        {len(no_match)}")
    print()

    if nears:
        print("  ⚠ ΚΟΝΤΙΝΑ ΤΑΙΡΙΑΣΜΑΤΑ — έλεγξέ τα:")
        for w in nears[:10]:
            print(f"      Sheet {w['date']:%d/%m/%Y}  ←  Excel {w['excel_date']:%d/%m/%Y}"
                  f"  →  {w['number']}")
        if len(nears) > 10:
            print(f"      …και άλλα {len(nears) - 10}")
        print()

    if no_match:
        print(f"  · Ημερομηνίες Sheet χωρίς αριθμό στο Excel (πρώτες 8):")
        for d in no_match[:8]:
            print(f"      {d:%d/%m/%Y}")
        if len(no_match) > 8:
            print(f"      …και άλλες {len(no_match) - 8}")
        print()

    # Προειδοποίηση αν τα ποσά διαφέρουν — μπορεί να είναι λάθος ταίριασμα.
    mismatched = [
        w for w in to_write
        if abs(w["sheet_amount"] - w["excel_amount"]) > 1.0
        and w["excel_amount"] > 0
    ]
    if mismatched:
        print(f"  ⚠ {len(mismatched)} γραμμές με διαφορετικό ΠΟΣΟ:")
        for w in mismatched[:8]:
            print(f"      {w['date']:%d/%m/%Y}  Sheet {w['sheet_amount']:>10,.2f} "
                  f"vs Excel {w['excel_amount']:>10,.2f}  → {w['number']}")
        if len(mismatched) > 8:
            print(f"      …και άλλες {len(mismatched) - 8}")
        print("      (Το ταίριασμα γίνεται με ΗΜΕΡΟΜΗΝΙΑ — έλεγξέ τα αν σου φαίνονται λάθος.)")
        print()

    if to_write:
        print("  Δείγμα από αυτά που θα γραφτούν:")
        for w in to_write[:10]:
            print(f"      γραμμή {w['row']:>4}  {w['date']:%d/%m/%Y}  →  {w['number']}")
        if len(to_write) > 10:
            print(f"      …και άλλα {len(to_write) - 10}")
        print()

    if args.dry:
        print("· ΔΟΚΙΜΗ — δεν γράφτηκε τίποτα.")
        print("  Τρέξε ξανά χωρίς --dry για να εφαρμοστεί.")
        return 0

    if not to_write:
        print("✓ Τίποτα να γραφτεί.")
        return 0

    # ── 5. ΓΡΑΨΕ ──
    print(f"▶ Εγγραφή {len(to_write)} αριθμών…")
    ok = fail = 0

    for i, w in enumerate(to_write, 1):
        if update_timologiseis_field(w["row"], "check_number", w["number"]):
            ok += 1
        else:
            fail += 1
            print(f"  ✗ Γραμμή {w['row']} ({w['date']:%d/%m/%Y})")

        # Το Sheets API αντέχει ~60 κλήσεις/λεπτό.
        time.sleep(1.1)

        if i % 10 == 0:
            print(f"  · {i}/{len(to_write)}")

    print()
    print(f"✓ Γράφτηκαν {ok}" + (f", απέτυχαν {fail}" if fail else ""))
    return 0 if not fail else 1


if __name__ == "__main__":
    sys.exit(main())
